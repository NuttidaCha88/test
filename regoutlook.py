import requests
import time
import random
import re
import threading
import string
import json
import os
import sys
import atexit
import signal
import base64
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys

# Bổ sung biến và locks mới để tránh nghẽn cổ chai
excel_io_lock = threading.Lock()  # Lock riêng cho thao tác Excel
log_io_lock = threading.Lock()    # Lock cho các hoạt động ghi log
api_lock = threading.Lock()       # Lock cho các API request (nếu cần)
recovery_email_lock = threading.Lock() # Lock cho việc quản lý email khôi phục

# Trạng thái các profile đang xử lý (để khôi phục khi bị tắt đột ngột)
processing_profiles = {}
processing_profiles_lock = threading.Lock()

# Trạng thái email khôi phục đang được sử dụng
used_recovery_emails = set()
using_recovery_emails = {}  # thread_id -> email

# Biến để theo dõi trạng thái chương trình
is_shutting_down = False

# ---------------------------------------
# Hàm ghi log thread-safe
# ---------------------------------------
def write_to_log(filename, content):
    """Hàm ghi log thread-safe"""
    with log_io_lock:
        try:
            with open(filename, "a") as f:
                f.write(content + "\n")
            return True
        except Exception as e:
            print(f"Lỗi khi ghi vào file {filename}: {e}")
            return False

# ---------------------------------------
# Hàm lưu Excel với khả năng retry
# ---------------------------------------
def save_excel_with_retry(max_retries=3, wait_between_retries=1, create_backup=False):
    """
    Lưu file Excel với khả năng thử lại nếu thất bại
    Sử dụng cơ chế lock file tạm thời để tránh xung đột
    
    Args:
        max_retries: Số lần thử lại tối đa
        wait_between_retries: Thời gian đợi giữa các lần thử lại (giây)
        create_backup: Có tạo file backup không (True/False)
    """
    with excel_io_lock:
        # Tạo file lock để chỉ ra rằng Excel đang được lưu
        lock_file = 'profiles.xlsx.lock'
        try:
            # Kiểm tra xem đã có tiến trình nào đang lưu file không
            if os.path.exists(lock_file):
                lock_time = os.path.getmtime(lock_file)
                current_time = time.time()
                # Nếu file lock quá cũ (>30 giây), có thể tiến trình trước đã bị crash
                if current_time - lock_time > 30:
                    print(f"Phát hiện file lock cũ (>30s), có thể tiến trình trước đã crash. Xóa và tiếp tục.")
                    os.remove(lock_file)
                else:
                    print(f"Phát hiện file đang được lưu bởi tiến trình khác. Đợi 1 giây và thử lại.")
                    time.sleep(1)
                    return save_excel_with_retry(max_retries, wait_between_retries, create_backup)  # Thử lại
            
            # Tạo file lock trước khi lưu
            with open(lock_file, 'w') as f:
                f.write(f"Lock created at {time.ctime()}")
            
            # Tạo một backup duy nhất nếu được yêu cầu
            if create_backup and os.path.exists('profiles.xlsx'):
                backup_file = 'profiles_backup.xlsx'
                try:
                    import shutil
                    shutil.copy2('profiles.xlsx', backup_file)
                    print(f"Đã tạo file backup: {backup_file}")
                except Exception as backup_error:
                    print(f"Không thể tạo backup: {backup_error}")
            
            temp_file = 'profiles_temp.xlsx'
            for attempt in range(max_retries):
                try:
                    # Lưu vào file tạm thời trước
                    workbook.save(temp_file)
                    
                    # Kiểm tra tính toàn vẹn của file tạm thời
                    try:
                        # Thử mở file để xác nhận nó không bị hỏng
                        temp_wb = load_workbook(temp_file)
                        temp_wb.close()
                    except Exception as integrity_error:
                        print(f"File tạm thời bị hỏng: {integrity_error}")
                        # Nếu file tạm thời bị hỏng, xóa nó và thử lại
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                        raise integrity_error  # Ném lỗi để thử lại
                    
                    # Thay thế file chính bằng file tạm thời
                    try:
                        import shutil
                        shutil.move(temp_file, 'profiles.xlsx')
                    except Exception as move_error:
                        print(f"Lỗi khi thay thế file: {move_error}")
                        # Nếu không thể thay thế, thử lưu trực tiếp
                        workbook.save('profiles.xlsx')
                    
                    print(f"Đã lưu file Excel thành công (lần thử {attempt+1})")
                    return True
                except PermissionError as pe:
                    print(f"Lỗi quyền truy cập khi lưu Excel (lần thử {attempt+1}/{max_retries}): {pe}")
                    if attempt < max_retries - 1:
                        print(f"Đợi {wait_between_retries} giây và thử lại...")
                        time.sleep(wait_between_retries * 2)  # Tăng thời gian đợi
                    else:
                        # Nếu đã hết số lần thử, tạo backup khẩn cấp
                        if not os.path.exists('profiles_emergency_backup.xlsx'):
                            try:
                                workbook.save('profiles_emergency_backup.xlsx')
                                print(f"Đã lưu vào file backup khẩn cấp")
                            except Exception as backup_e:
                                print(f"Không thể lưu file backup khẩn cấp: {backup_e}")
                        return False
                except Exception as e:
                    print(f"Lỗi khi lưu Excel (lần thử {attempt+1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        print(f"Đợi {wait_between_retries} giây và thử lại...")
                        time.sleep(wait_between_retries)
                    else:
                        # Nếu đã hết số lần thử, tạo backup khẩn cấp
                        if not os.path.exists('profiles_emergency_backup.xlsx'):
                            try:
                                workbook.save('profiles_emergency_backup.xlsx')
                                print(f"Đã lưu vào file backup khẩn cấp")
                            except Exception as backup_e:
                                print(f"Không thể lưu file backup khẩn cấp: {backup_e}")
                        return False
            return False
        finally:
            # Đảm bảo luôn xóa file lock khi hoàn thành
            if os.path.exists(lock_file):
                try:
                    os.remove(lock_file)
                except:
                    pass

# ---------------------------------------
# Hàm theo dõi profile đang xử lý
# ---------------------------------------
def mark_profile_as_processing(thread_id, profile_id, row_number):
    """Đánh dấu profile đang được xử lý"""
    with processing_profiles_lock:
        processing_profiles[profile_id] = {
            "thread_id": thread_id,
            "row": row_number,
            "start_time": time.time()
        }

def mark_profile_as_completed(profile_id):
    """Đánh dấu profile đã xử lý xong"""
    with processing_profiles_lock:
        if profile_id in processing_profiles:
            del processing_profiles[profile_id]

# ---------------------------------------
# Hàm xử lý khi chương trình tắt
# ---------------------------------------
def save_state_on_exit():
    """Lưu trạng thái khi chương trình tắt (cả tắt bình thường và đột ngột)"""
    global is_shutting_down
    is_shutting_down = True
    print("Đang lưu trạng thái trước khi thoát...")
    
    # Lưu file Excel một lần cuối
    with excel_io_lock:
        try:
            # Tạo backup một lần khi thoát
            save_excel_with_retry(create_backup=True)
            print("Đã lưu file profiles.xlsx thành công")
        except Exception as e:
            print(f"Lỗi khi lưu file profiles.xlsx: {e}")
    
    # Lưu danh sách các profile đang xử lý
    with processing_profiles_lock:
        if processing_profiles:
            try:
                with open("processing_profiles.json", "w") as f:
                    json.dump(processing_profiles, f)
                print(f"Đã lưu trạng thái của {len(processing_profiles)} profile đang xử lý")
            except Exception as e:
                print(f"Lỗi khi lưu trạng thái profile: {e}")
                
    # Xóa các file tạm thời nếu có
    try:
        if os.path.exists('profiles_temp.xlsx'):
            os.remove('profiles_temp.xlsx')
        if os.path.exists('profiles.xlsx.lock'):
            os.remove('profiles.xlsx.lock')
    except Exception as e:
        print(f"Không thể xóa file tạm: {e}")

# ---------------------------------------
# Hàm khôi phục trạng thái trước đó
# ---------------------------------------
def recover_from_previous_run():
    """Khôi phục trạng thái từ lần chạy trước nếu bị ngắt đột ngột"""
    try:
        if os.path.exists("processing_profiles.json"):
            with open("processing_profiles.json", "r") as f:
                interrupted_profiles = json.load(f)
            
            if interrupted_profiles:
                print(f"Phát hiện {len(interrupted_profiles)} profile bị gián đoạn từ lần chạy trước")
                
                # Đánh dấu các profile này là "Bị gián đoạn" trong Excel
                with excel_io_lock:
                    for profile_id, info in interrupted_profiles.items():
                        row = info.get("row")
                        if row:
                            try:
                                worksheet.cell(row=row, column=10).value = "Bị gián đoạn"
                                print(f"Đã đánh dấu profile {profile_id} (hàng {row}) là 'Bị gián đoạn'")
                            except:
                                print(f"Không thể đánh dấu profile {profile_id}")
                    
                    # Lưu thay đổi
                    save_excel_with_retry()
                
                # Xóa file trạng thái
                os.remove("processing_profiles.json")
                print("Đã xóa file trạng thái cũ")
            
    except Exception as e:
        print(f"Lỗi khi khôi phục trạng thái: {e}")

# Đăng ký hàm xử lý khi thoát
atexit.register(save_state_on_exit)

# Đăng ký xử lý tín hiệu
def signal_handler(sig, frame):
    print(f"Đã nhận tín hiệu {sig}, đang chuẩn bị thoát...")
    save_state_on_exit()
    sys.exit(0)

# Đăng ký xử lý các tín hiệu
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # kill
if hasattr(signal, 'SIGBREAK'):  # Windows
    signal.signal(signal.SIGBREAK, signal_handler)

# ---------------------------------------
# Hàm lấy proxy từ API key
# ---------------------------------------
def get_proxy_from_api(api_key):
    """
    Lấy proxy từ API key thông qua API WWProxy
    Trả về chuỗi proxy nếu thành công, None nếu thất bại, hoặc dict với key "wait" nếu cần đợi
    """
    api_url = f"https://wwproxy.com/api/client/proxy/available?key={api_key}&provinceId=-1"
    
    try:
        response = requests.get(api_url)
        data = response.json()
        
        if data.get("status") == "OK" and data.get("data") and data["data"].get("proxy"):
            proxy = data["data"]["proxy"]
            print(f"Đã lấy được proxy: {proxy} từ API key: {api_key}")
            return proxy
        elif data.get("status") == "BAD_REQUEST" and "Vui lòng chờ thêm" in data.get("message", ""):
            # Trích xuất thời gian cần chờ từ thông báo lỗi
            wait_time_match = re.search(r'(\d+)s', data.get("message", ""))
            if wait_time_match:
                wait_seconds = int(wait_time_match.group(1)) + 2
                print(f"Cần đợi {wait_seconds} giây trước khi lấy proxy mới từ API key: {api_key}")
                return {"wait": wait_seconds}
        
        print(f"Lỗi khi lấy proxy từ API key {api_key}: {data}")
        return None
    except Exception as e:
        print(f"Exception khi lấy proxy từ API key {api_key}: {e}")
        return None

# ---------------------------------------
# Hàm cập nhật proxy cho profile
# ---------------------------------------
def update_proxy(profile_id, raw_proxy):
    """
    Cập nhật proxy cho profile trước khi mở.
    Gửi POST request với raw_proxy.
    Nếu trả về "Profile not found" thì log profile_id ra file profileloi.txt.
    """
    update_url = f"http://127.0.0.1:19995/api/v3/profiles/update/{profile_id}"
    headers = {"accept": "application/json", "Content-Type": "application/json"}
    data = {"raw_proxy": f"{raw_proxy}"} # Không cần thêm http:// prefix
    try:
        r = requests.post(update_url, headers=headers, json=data)
        r.raise_for_status()
        response_json = r.json()
        if response_json.get("success"):
            print(f"Proxy updated successfully for profile {profile_id}.")
            return True
        elif response_json.get("message") == "Profile not found":
            print(f"Update failed. Profile not found: {profile_id}")
            with open("profileloi.txt", "a") as f:
                f.write(str(profile_id) + "\n")
            return False
        else:
            print(f"Unexpected response when updating proxy for profile {profile_id}: {response_json}")
            return False
    except Exception as e:
        print(f"Exception updating proxy for profile {profile_id}: {e}")
        return False

# ---------------------------------------
# Hàm tạo tên ngẫu nhiên
# ---------------------------------------
def generate_random_name():
    """
    Tạo tên người dùng ngẫu nhiên kiểu tên tiếng Anh
    """
    first_names = [
        "James", "John", "Robert", "Michael", "William", "David", "Richard", "Joseph", "Thomas", "Charles",
        "Christopher", "Daniel", "Matthew", "Anthony", "Mark", "Donald", "Steven", "Paul", "Andrew", "Joshua",
        "Kenneth", "Kevin", "Brian", "George", "Timothy", "Emma", "Olivia", "Ava", "Isabella", "Sophia",
        "Charlotte", "Mia", "Amelia", "Harper", "Evelyn", "Abigail", "Emily", "Elizabeth", "Mila", "Ella",
        "Avery", "Sofia", "Camila", "Aria", "Scarlett", "Jonathan", "Nathan", "Oliver", "Henry", "Sebastian",
        "Alexander", "Benjamin", "Samuel", "Patrick", "Nicholas", "Gregory", "Ryan", "Brandon", "Adam", "Zachary",
        "Justin", "Jose", "Ethan", "Jacob", "Tyler", "Austin", "Jordan", "Madison", "Chloe", "Penelope",
        "Lily", "Layla", "Grace", "Zoey", "Nora", "Riley", "Addison", "Eleanor", "Claire", "Stella", "Violet",
        "Gabriel", "Logan", "Owen", "Lucas", "Carter", "Caleb", "Hunter", "Luke", "Landon", "Connor",
        "Aiden", "Jack", "Wyatt", "Jayden", "Dylan", "Cole", "Dominic", "Levi", "Brody", "Max",
        "Liam", "Asher", "Owen", "Declan", "Theodore", "Hazel", "Aurora", "Willow", "Luna", "Nova",
        "Skylar", "Genesis", "Naomi", "Leah", "Audrey", "Brooklyn", "Paisley", "Savannah", "Kennedy", "Allison",
        "Maya", "Valentina", "Autumn", "Summer", "Winter", "Violet", "Ruby", "Jade", "Hazel", "Eleanor",
        "Penelope", "Athena", "Alice", "June", "Sadie", "Ophelia", "Eliza", "Hazel", "Willow", "Juniper",
        "Iris", "Florence", "Rosalie", "Adeline", "Maisie", "Eloise", "Cecilia", "Clara", "Josephine", "Rose",
        "Arthur", "Leo", "Milo", "Jasper", "Hugo", "August", "Finn", "Miles", "Silas", "Owen", "Rhys",
        "Felix", "Otis", "Elijah", "Cyrus", "Ezra", "Atticus", "Oscar", "Gideon", "Remi", "Leo", "Cody",
        "Jake", "Toby", "Jared", "Felix", "An", "Binh", "Cuong", "Dung", "Duc", "Giang", "Hai", "Hang", "Hieu", "Hoa",
        "Hong", "Hue", "Huong", "Lan", "Linh", "Long", "Mai", "Nam", "Nga", "Ngoc",
        "Oanh", "Phuong", "Quan", "Quynh", "Sang", "Son", "Thanh", "Thao", "Thi", "Thuy",
        "Trang", "Tu", "Tung", "Van", "Viet", "Yen", "Minh", "Khanh", "Thang", "Tuan",
        "Anh", "Diep", "Uyen", "Vy", "Hoang", "Nguyen", "Phong", "Kim", "Thuy", "Giau",
        "Loc", "Phuoc", "Tai", "Thien", "Nhan", "Tri", "Tin", "Nghia", "Nhat", "Nguyet"
    ]
    
    last_names = [
        "Smith", "Johnson", "Williams", "Jones", "Brown", "Davis", "Miller", "Wilson", "Moore", "Taylor",
        "Anderson", "Thomas", "Jackson", "White", "Harris", "Martin", "Thompson", "Garcia", "Martinez", "Robinson",
        "Clark", "Rodriguez", "Lewis", "Lee", "Walker", "Hall", "Allen", "Young", "Hernandez", "King",
        "Wright", "Lopez", "Hill", "Scott", "Green", "Adams", "Baker", "Gonzalez", "Nelson", "Carter",
        "Mitchell", "Perez", "Roberts", "Turner", "Phillips", "Campbell", "Parker", "Evans", "Edwards", "Collins",
        "Stewart", "Morris", "Murphy", "Cook", "Rogers", "Morgan", "Peterson", "Cooper", "Reed", "Bailey",
        "Bell", "Sullivan", "Jenkins", "Perry", "Powell", "Long", "Gray", "Ross", "Hughes", "Wood",
        "Myers", "Bennett", "Ward", "Florence", "Henderson", "Simmons", "Webb", "Fisher", "Holmes", "Ford",
        "Shaw", "Rice", "Richards", "Davidson", "Burton", "Spencer", "Harvey", "Arnold", "Ryan", "Nichols",
        "Jordan", "Stone", "Bowman", "Frazier", "Gardner", "Newman", "Hicks", "Lowe", "Mueller", "Boyd",
        "Potter", "Chase", "Woodward", "Dennis", "West", "Horton", "Rose", "Logan", "Hunt", "Blackwell",
        "Carr", "Henry", "Gross", "McDonald", "Franklin", "Bishop", "Weaver", "Bush", "Bowers", "Fuller",
        "Jenkins", "Simmons", "Atkins", "Crawford", "Dennis", "Norman", "Hammond", "Saunders", "Moss",
        "Pittman", "Wallace", "Gilbert", "Glover", "Sutton", "McDaniel", "Lambert", "Townsend", "Fletcher",
        "Nicholson", "Lyons", "Mccarthy", "Sherman", "Valentine", "Blackburn", "Armstrong", "Bradshaw", "Crosby",
        "Rich", "Lang", "Winters", "Brewer", "Forbes", "Erickson", "Norton", "Brewer", "Lane", "Romero",
        "Vaughn", "Curtis", "Osborn", "Higgins", "Kaufman", "Morse", "Meyers", "Baldwin", "Warner",
        "Nguyen", "Tran", "Le", "Pham", "Huynh", "Phan", "Vo", "Dang", "Bui", "Do",
        "Ho", "Ngo", "Duong", "Dinh", "Cao", "Truong", "Mai", "Ton", "Luong", "Vuong",
        "Ta", "Thi", "Quach", "Ha", "La", "Banh", "Lieu", "Thai", "Ong", "Ma",
        "Kieu", "Ngan", "Chau", "Chiem", "Quang", "Diep", "Thach", "Kim", "Hua", "Vu",
        "Dam", "Bach", "Hang", "Tieu", "Phung", "Tu", "Loan", "Hoang", "Sam", "Y"
    ]
    
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    
    return first_name, last_name

# ---------------------------------------
# Hàm tạo mật khẩu ngẫu nhiên
# ---------------------------------------
def generate_secure_password():
    """
    Tạo mật khẩu ngẫu nhiên với 10 ký tự có ít nhất 1 chữ hoa, 1 chữ thường, 1 số, 1 ký tự đặc biệt
    """
    uppercase_chars = string.ascii_uppercase
    lowercase_chars = string.ascii_lowercase
    digits = string.digits
    special_chars = "!@#$%^&*()-_=+"
    
    # Đảm bảo có ít nhất 1 ký tự từ mỗi loại
    password = [
        random.choice(uppercase_chars),
        random.choice(lowercase_chars),
        random.choice(digits),
        random.choice(special_chars)
    ]
    
    # Thêm 6 ký tự ngẫu nhiên nữa để đủ 10 ký tự
    all_chars = uppercase_chars + lowercase_chars + digits + special_chars
    password.extend(random.choice(all_chars) for _ in range(6))
    
    # Trộn ngẫu nhiên các ký tự trong mật khẩu
    random.shuffle(password)
    
    return ''.join(password)

# ---------------------------------------
# Hàm tạo ngày tháng năm sinh ngẫu nhiên
# ---------------------------------------
def generate_birth_date():
    """
    Tạo ngày tháng năm sinh ngẫu nhiên để người dùng trong độ tuổi 20-30
    """
    current_year = 2025  # Đặt năm hiện tại là 2025
    
    # Người dùng trong độ tuổi 20-30
    birth_year = random.randint(current_year - 30, current_year - 20)
    birth_month = random.randint(1, 12)
    
    # Xác định số ngày tối đa trong tháng
    max_days = 31
    if birth_month in [4, 6, 9, 11]:
        max_days = 30
    elif birth_month == 2:
        # Kiểm tra năm nhuận
        if (birth_year % 400 == 0) or (birth_year % 4 == 0 and birth_year % 100 != 0):
            max_days = 29
        else:
            max_days = 28
            
    birth_day = random.randint(1, max_days)
    
    return birth_day, birth_month, birth_year

# ---------------------------------------
# Đọc file proxy.txt - Đọc API keys
# ---------------------------------------
try:
    with open("proxy.txt", "r") as f:
        api_keys = [line.strip() for line in f.readlines()]
    
    if len(api_keys) < 7:
        print(f"Cảnh báo: Có ít hơn 6 API key trong file proxy.txt. Một số luồng sẽ không hoạt động.")
    
    print(f"Đã tải {len(api_keys)} API key từ file proxy.txt")

except Exception as e:
    print(f"Error reading proxy.txt: {e}")
    api_keys = []

# ---------------------------------------
# Đọc file Excel profiles.xlsx
#  - Cột A: Profile ID
#  - Cột J (cột thứ 10): Kết quả (kiểm tra KHÔNG phải "Thành Công")
# ---------------------------------------
workbook = load_workbook('profiles.xlsx')
worksheet = workbook.active

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
# Đặt định dạng default_fill thành None (không có màu nền)
default_fill = PatternFill(fill_type=None)

profiles = []
for i, row in enumerate(worksheet.iter_rows(min_row=2, max_col=10, values_only=False), start=2):
    profile_id = row[0].value  # Cột A
    result_value = worksheet.cell(row=i, column=10).value  # Cột J
    
    # Chỉ xử lý nếu cột J KHÔNG phải "Thành Công"
    if not result_value or str(result_value).strip() != "Thành Công":
        profiles.append({
            "id": profile_id,
            "row": i
        })
    else:
        print(f"Skipping profile {profile_id} (row {i}) vì cột J đã là 'Thành Công'.")

# Biến toàn cục để xử lý đa luồng
profiles_lock = threading.Lock()
profile_index = 0  # index profile hiện tại

# ---------------------------------------
# Hàm lấy refresh token từ localhost:5000
# ---------------------------------------
def get_refresh_token(driver, wait, username, password, thread_id, profile_id, row_number, retries=3):
    """
    Truy cập localhost:5000 và lấy refresh token
    Kiểm tra và click vào newSessionLink nếu có trước khi đợi animation xuất hiện
    
    Args:
        driver: WebDriver instance
        wait: WebDriverWait instance
        username: Tên người dùng
        password: Mật khẩu
        thread_id: ID của thread đang xử lý
        profile_id: ID của profile đang xử lý
        row_number: Số thứ tự hàng trong Excel
        retries: Số lần thử lại nếu thất bại
        
    Returns:
        refresh_token nếu thành công, None nếu thất bại
    """
    for attempt in range(retries):
        try:
            # Truy cập https://tolive.site - sẽ tự động chuyển hướng đến Microsoft
            driver.get("https://tolive.site")
            print(f"Thread {thread_id}: Truy cập https://tolive.site, đang chờ chuyển hướng...")
            time.sleep(5)
            # Đợi trang load xong trong 60 giây
            try:
                WebDriverWait(driver, 60).until(lambda d: d.execute_script("return document.readyState") == "complete")
                print(f"Thread {thread_id}: Trang tolive.site đã load xong")
            except TimeoutException:
                print(f"Thread {thread_id}: Timeout sau 60 giây chờ trang tolive.site load")
                return None
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi chờ trang load: {e}")
            
            # Kiểm tra xem tài khoản có bị khóa không TRƯỚC khi kiểm tra newSessionLink
            try:
                account_locked_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@role='heading' and contains(text(), 'Tài khoản của bạn đã bị khóa') or contains(text(), 'Your account has been suspended')]"))
                )
                print(f"Thread {thread_id}: [CẢNH BÁO] Phát hiện thông báo tài khoản bị khóa")
                
                
                # Đánh dấu là tài khoản bị khóa trong Excel
                with profiles_lock:
                    worksheet.cell(row=row_number, column=10).value = "Tài khoản bị khóa"
                    # Tô đỏ dòng đó
                    for col in range(1, 12):  # Bao gồm cả cột K (cột thứ 11)
                        worksheet.cell(row=row_number, column=col).fill = red_fill
                    # Lưu file Excel
                    save_excel_with_retry()
                    print(f"Thread {thread_id}: Đã đánh dấu 'Tài khoản bị khóa' cho profile {profile_id}")
                
                # Đóng profile
                close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                try:
                    requests.get(close_url)
                    print(f"Thread {thread_id}: Đã đóng profile {profile_id} do tài khoản bị khóa")
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi đóng profile {profile_id}: {e}")
                
                return "ACCOUNT_LOCKED"
                
            except TimeoutException:
                print(f"Thread {thread_id}: Không phát hiện thông báo tài khoản bị khóa trước bước kiểm tra newSessionLink")
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi kiểm tra tài khoản bị khóa trước bước newSessionLink: {e}")
            
            # Kiểm tra xem có phần tử newSessionLink không
            try:
                new_session_link = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "newSessionLink"))
                )
                print(f"Thread {thread_id}: Phát hiện phần tử newSessionLink (tài khoản đã đăng nhập trước đó), đang click vào...")
                
                # Lấy email hiển thị từ phần tử để ghi log
                try:
                    email_display = driver.find_element(By.XPATH, "//div[@id='newSessionLink']//small[contains(text(), '@')]").text
                    print(f"Thread {thread_id}: Tài khoản đã đăng nhập: {email_display}")
                except:
                    print(f"Thread {thread_id}: Không thể lấy thông tin email từ tài khoản đã đăng nhập")
                
                # Click vào phần tử
                simulate_human_click(driver, new_session_link)
                print(f"Thread {thread_id}: Đã click vào newSessionLink")
            except TimeoutException:
                print(f"Thread {thread_id}: Không tìm thấy phần tử newSessionLink, có thể đây là lần đăng nhập đầu tiên")
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi xử lý newSessionLink: {e}")
            
            # Bổ sung thêm bước kiểm tra tài khoản bị khóa
            try:
                account_locked_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@role='heading' and contains(text(), 'Tài khoản của bạn đã bị khóa') or contains(text(), 'Your account has been suspended')]"))
                )
                print(f"Thread {thread_id}: [CẢNH BÁO] Phát hiện thông báo tài khoản bị khóa")
                
                
                # Đánh dấu là tài khoản bị khóa trong Excel
                with profiles_lock:
                    worksheet.cell(row=row_number, column=10).value = "Tài khoản bị khóa"
                    # Tô đỏ dòng đó
                    for col in range(1, 12):  # Bao gồm cả cột K (cột thứ 11)
                        worksheet.cell(row=row_number, column=col).fill = red_fill
                    # Lưu file Excel
                    save_excel_with_retry()
                    print(f"Thread {thread_id}: Đã đánh dấu 'Tài khoản bị khóa' cho profile {profile_id}")
                
                # Đóng profile
                close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                try:
                    requests.get(close_url)
                    print(f"Thread {thread_id}: Đã đóng profile {profile_id} do tài khoản bị khóa")
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi đóng profile {profile_id}: {e}")
                
                return "ACCOUNT_LOCKED"
                
            except TimeoutException:
                print(f"Thread {thread_id}: Không phát hiện thông báo tài khoản bị khóa, tiếp tục quy trình")
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi kiểm tra tài khoản bị khóa: {e}")
            
            # Kiểm tra xem tài khoản có bị khóa không
            try:
                account_locked_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@role='heading' and @aria-level='1' and @id='serviceAbuseLandingTitle' and contains(text(), 'Tài khoản của bạn đã bị khóa')]"))
                )
                print(f"Thread {thread_id}: [CẢNH BÁO] Phát hiện thông báo 'Tài khoản của bạn đã bị khóa'")
                
                # Đánh dấu là tài khoản bị khóa trong Excel
                with profiles_lock:
                    worksheet.cell(row=row_number, column=10).value = "Tài khoản bị khóa"
                    # Tô đỏ dòng đó
                    for col in range(1, 11):
                        worksheet.cell(row=row_number, column=col).fill = red_fill
                    # Lưu file Excel
                    save_excel_with_retry()
                    print(f"Thread {thread_id}: Đã đánh dấu 'Tài khoản bị khóa' cho profile {profile_id}")
                
                # Đóng profile
                close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                try:
                    requests.get(close_url)
                    print(f"Thread {thread_id}: Đã đóng profile {profile_id} do tài khoản bị khóa")
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi đóng profile {profile_id}: {e}")
                
                return "ACCOUNT_LOCKED"
                
            except TimeoutException:
                print(f"Thread {thread_id}: Không phát hiện thông báo 'Tài khoản của bạn đã bị khóa', tiếp tục quy trình")
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi kiểm tra tài khoản bị khóa: {e}")
            
            # Đợi phần tử animation xuất hiện trong 60 giây
            try:
                print(f"Thread {thread_id}: Đang chờ phần tử animation xuất hiện (tối đa 60 giây)...")
                route_animation = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='routeAnimation']"))
                )
                print(f"Thread {thread_id}: Đã phát hiện phần tử animation loading, đang đợi nút consent xuất hiện...")
                
                # Đợi nút consent xuất hiện và click
                try:
                    consent_button = WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='appConsentPrimaryButton']"))
                    )
                    # Sử dụng hàm click người dùng thật
                    simulate_human_click(driver, consent_button)
                    print(f"Thread {thread_id}: Đã nhấp vào nút Consent/Đồng ý")
                    
                    # Đợi trang load xong và chuyển hướng đến trang kết quả
                    WebDriverWait(driver, 60).until(
                        lambda d: "getAToken" in d.current_url
                    )
                    print(f"Thread {thread_id}: Đã chuyển hướng đến trang kết quả")
                    #time.sleep(3)  # Đợi trang tải hoàn tất
                    
                    # Lấy refresh token từ nội dung trang
                    page_content = driver.page_source
                    refresh_token_match = re.search(r'<h3>Refresh Token:</h3>\s*<p>(.*?)</p>', page_content, re.DOTALL)
                    
                    if refresh_token_match:
                        refresh_token = refresh_token_match.group(1).strip()
                        print(f"Thread {thread_id}: Đã lấy được refresh token: {refresh_token[:20]}...")
                        return refresh_token
                    else:
                        # Thử lấy từ localStorage nếu không tìm thấy trong HTML
                        try:
                            tokens_json = driver.execute_script("return localStorage.getItem('tokens');")
                            if tokens_json:
                                tokens = json.loads(tokens_json)
                                if 'refresh_token' in tokens:
                                    print(f"Thread {thread_id}: Đã lấy được refresh token từ localStorage")
                                    return tokens['refresh_token']
                        except Exception as e:
                            print(f"Thread {thread_id}: Lỗi khi lấy token từ localStorage: {e}")
                    
                    print(f"Thread {thread_id}: Không tìm thấy refresh token sau khi click nút consent")
                    return None
                    
                except TimeoutException:
                    print(f"Thread {thread_id}: Không tìm thấy nút consent sau khi đợi")
                    return None
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi xử lý nút consent: {e}")
                    return None
                    
            except TimeoutException:
                print(f"Thread {thread_id}: [LỖI] Không tìm thấy phần tử animation sau khi đợi 60 giây")
                
                # Đánh dấu là lỗi token trong Excel
                with profiles_lock:
                    worksheet.cell(row=row_number, column=10).value = "Lỗi Token - Không tìm thấy animation"
                    # Tô đỏ dòng đó
                    for col in range(1, 11):
                        worksheet.cell(row=row_number, column=col).fill = red_fill
                    # Lưu file Excel
                    save_excel_with_retry()
                    print(f"Thread {thread_id}: Đã đánh dấu 'Lỗi Token - Không tìm thấy animation' cho profile {profile_id}")
                
                # Đóng profile
                close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                try:
                    requests.get(close_url)
                    print(f"Thread {thread_id}: Đã đóng profile {profile_id} do không tìm thấy animation sau 60 giây")
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi đóng profile {profile_id}: {e}")
                
                return "FAILED_ANIMATION"
                
            except Exception as e:
                print(f"Thread {thread_id}: Lỗi khi đợi phần tử animation: {e}")
                return None
                
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi khi lấy refresh token, lần thử {attempt+1}: {e}")
            #time.sleep(5)
    
    return None

# ---------------------------------------
# Hàm mô phỏng đánh máy như người thật
# ---------------------------------------
def simulate_human_typing(element, text, min_delay=0.05, max_delay=0.15):
    """
    Nhập văn bản vào element một cách chậm rãi như người thật
    với độ trễ ngẫu nhiên giữa các ký tự
    
    Args:
        element: WebElement để nhập văn bản
        text: Chuỗi văn bản cần nhập
        min_delay: Thời gian trễ tối thiểu giữa các ký tự (giây)
        max_delay: Thời gian trễ tối đa giữa các ký tự (giây)
    """
    for char in text:
        element.send_keys(char)
        # Độ trễ ngẫu nhiên giữa các ký tự
        time.sleep(random.uniform(min_delay, max_delay))
    
    # Đợi một chút sau khi nhập xong
    time.sleep(random.uniform(0.3, 0.7))

# ---------------------------------------
# Hàm mô phỏng hành vi click chuột người thật
# ---------------------------------------
def simulate_human_click(driver, element):
    """
    Thực hiện click với độ trễ ngẫu nhiên trước và sau khi click
    
    Args:
        driver: WebDriver instance
        element: WebElement để click
    """
    # Đợi một khoảng thời gian ngẫu nhiên trước khi click
    time.sleep(random.uniform(0.3, 1.2))
    
    # Thử sử dụng JavaScript để rê chuột vào phần tử (mô phỏng hover)
    try:
        driver.execute_script("""
            var event = new MouseEvent('mouseover', {
                'view': window,
                'bubbles': true,
                'cancelable': true
            });
            arguments[0].dispatchEvent(event);
        """, element)
        time.sleep(random.uniform(0.1, 0.3))  # Đợi một chút sau khi hover
    except:
        pass  # Bỏ qua nếu không thành công
    
    # Thực hiện click
    element.click()
    
    # Đợi một khoảng thời gian ngẫu nhiên sau khi click
    time.sleep(random.uniform(0.2, 0.8))

# ---------------------------------------
# Hàm lấy email khôi phục từ danh sách
# ---------------------------------------
def get_recovery_email(thread_id):
    """
    Chọn ngẫu nhiên một email khôi phục từ file recovery_mail.txt
    và đảm bảo không có thread nào khác đang sử dụng email này
    
    Returns:
        tuple: (email, client_id, client_secret, refresh_token) hoặc None nếu không tìm được email
    """
    with recovery_email_lock:
        try:
            # Đọc danh sách email khôi phục
            recovery_emails = []
            with open("recovery_mail.txt", "r") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#"):
                        parts = line.split("|")
                        if len(parts) == 4:  # email|client_id|client_secret|refresh_token
                            recovery_emails.append(tuple(parts))
            
            # Nếu không có email nào, trả về None
            if not recovery_emails:
                print(f"Thread {thread_id}: Không có email khôi phục nào trong file recovery_mail.txt")
                return None
                
            # Đảm bảo chúng ta có ít nhất 12 email (để mỗi thread có 1 email khác nhau)
            if len(recovery_emails) < num_threads:
                print(f"Thread {thread_id}: Cảnh báo! Chỉ có {len(recovery_emails)} email khôi phục, nhưng có {num_threads} thread đang chạy.")
            
            # Lọc ra những email chưa được sử dụng
            available_emails = [email_data for email_data in recovery_emails if email_data[0] not in used_recovery_emails]
            
            # Nếu tất cả email đều đã được sử dụng, thử tìm email không được sử dụng bởi bất kỳ thread nào đang chạy
            if not available_emails:
                print(f"Thread {thread_id}: Tất cả email đã được sử dụng ít nhất một lần, đang tìm email có thể tái sử dụng...")
                # Lấy danh sách email đang được sử dụng bởi các thread
                currently_in_use = set(using_recovery_emails.values())
                available_emails = [email_data for email_data in recovery_emails if email_data[0] not in currently_in_use]
                
                # Nếu vẫn không có email nào khả dụng, đợi 5 giây và trả về None
                if not available_emails:
                    print(f"Thread {thread_id}: Không có email khôi phục nào khả dụng, đợi 5 giây và thử lại")
                    time.sleep(5)
                    return None
            
            # Chọn ngẫu nhiên một email khả dụng
            selected_email_data = random.choice(available_emails)
            email, client_id, client_secret, refresh_token = selected_email_data
            
            # Đánh dấu email này đã được sử dụng
            used_recovery_emails.add(email)
            using_recovery_emails[thread_id] = email
            
            print(f"Thread {thread_id}: Đã chọn email khôi phục: {email}")
            return selected_email_data
            
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi khi lấy email khôi phục: {e}")
            return None

# ---------------------------------------
# Hàm giải phóng email khôi phục
# ---------------------------------------
def release_recovery_email(thread_id):
    """Giải phóng email khôi phục để các thread khác có thể sử dụng"""
    with recovery_email_lock:
        if thread_id in using_recovery_emails:
            email = using_recovery_emails[thread_id]
            del using_recovery_emails[thread_id]
            print(f"Thread {thread_id}: Đã giải phóng email khôi phục {email}")

# ---------------------------------------
# Hàm lấy mã xác minh từ Gmail
# ---------------------------------------
def get_verification_code_from_gmail(thread_id, outlook_email, email, client_id, client_secret, refresh_token, max_retries=3):
    """
    Lấy mã xác minh từ Gmail bằng cách sử dụng Refresh Token
    
    Args:
        thread_id: ID của thread
        outlook_email: Email Outlook đang đăng ký (dạng username@outlook.com)
        email: Gmail để đọc mã
        client_id: Client ID cho OAuth2
        client_secret: Client Secret cho OAuth2
        refresh_token: Refresh Token để truy cập Gmail
        max_retries: Số lần thử lại tối đa
        
    Returns:
        Mã xác minh nếu thành công, None nếu thất bại
    """
    print(f"Thread {thread_id}: Đang lấy mã xác minh từ Gmail {email} cho tài khoản {outlook_email}")
    
    # Lấy username từ email Outlook (không bao gồm @outlook.com)
    outlook_username = outlook_email.split("@")[0]
    # Lấy 2 ký tự đầu và số cuối cùng của username
    username_first_chars = outlook_username[:2].lower()
    username_last_digit = None
    last_digit_match = re.search(r'(\d+)$', outlook_username)
    if last_digit_match:
        username_last_digit = last_digit_match.group(1)[-1]
    print(f"Thread {thread_id}: Pattern cần tìm trong email: {username_first_chars}**{username_last_digit if username_last_digit else ''}@outlook.com")
    
    for attempt in range(max_retries):
        try:
            # Lấy Access Token từ Refresh Token
            token_url = "https://oauth2.googleapis.com/token"
            token_data = {
                'refresh_token': refresh_token,
                'client_id': client_id,
                'client_secret': client_secret,
                'grant_type': 'refresh_token'
            }
            
            token_response = requests.post(token_url, data=token_data)
            if token_response.status_code != 200:
                print(f"Thread {thread_id}: Lỗi khi lấy access token: {token_response.text}")
                time.sleep(2)
                continue
            
            access_token = token_response.json().get('access_token')
            
            # Lấy danh sách email chưa đọc
            gmail_url = "https://www.googleapis.com/gmail/v1/users/me/messages"
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            params = {
                'q': 'is:unread from:account-security-noreply@accountprotection.microsoft.com subject:"Microsoft account security code"',
                'maxResults': 10
            }
            
            response = requests.get(gmail_url, headers=headers, params=params)
            if response.status_code != 200:
                print(f"Thread {thread_id}: Lỗi khi lấy danh sách email: {response.text}")
                time.sleep(2)
                continue
            
            messages_data = response.json()
            messages = messages_data.get('messages', [])
            
            if not messages:
                print(f"Thread {thread_id}: Không tìm thấy email chứa mã xác minh, đợi 5 giây và thử lại")
                time.sleep(5)
                continue
            
            # Lặp qua các email và tìm mã
            for msg in messages:
                msg_id = msg['id']
                msg_url = f"{gmail_url}/{msg_id}"
                msg_response = requests.get(msg_url, headers=headers)
                
                if msg_response.status_code != 200:
                    print(f"Thread {thread_id}: Lỗi khi lấy nội dung email: {msg_response.text}")
                    continue
                
                msg_data = msg_response.json()
                payload = msg_data.get('payload', {})
                
                # Lấy nội dung email
                body = ""
                if 'parts' in payload:
                    for part in payload['parts']:
                        if part['mimeType'] == 'text/plain':
                            if 'data' in part['body']:
                                body += base64.urlsafe_b64decode(part['body']['data']).decode('utf-8')
                elif 'body' in payload and 'data' in payload['body']:
                    body += base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8')
                
                # Kiểm tra xem email có chứa pattern cần tìm không
                masked_email_pattern = f"{username_first_chars}**{username_last_digit}@outlook.com" if username_last_digit else f"{username_first_chars}**@outlook.com"
                if masked_email_pattern.lower() not in body.lower():
                    print(f"Thread {thread_id}: Email không khớp với pattern {masked_email_pattern}, bỏ qua")
                    continue
                
                # Tìm mã xác minh
                security_code_match = re.search(r'Security code[:\s]*(\d{6})', body)
                if security_code_match:
                    verification_code = security_code_match.group(1)
                    print(f"Thread {thread_id}: Đã tìm thấy mã xác minh: {verification_code}")
                    
                    # Đánh dấu email là đã đọc
                    mark_as_read_url = f"{gmail_url}/{msg_id}/modify"
                    mark_data = {
                        'removeLabelIds': ['UNREAD']
                    }
                    requests.post(mark_as_read_url, headers=headers, json=mark_data)
                    
                    return verification_code
            
            print(f"Thread {thread_id}: Không tìm thấy mã xác minh trong các email, đợi 5 giây và thử lại (lần {attempt+1}/{max_retries})")
            time.sleep(5)
            
        except Exception as e:
            print(f"Thread {thread_id}: Lỗi khi lấy mã xác minh từ Gmail: {e}")
            time.sleep(2)
    
    print(f"Thread {thread_id}: Không thể lấy mã xác minh sau {max_retries} lần thử")
    return None

def process_profile(thread_id, api_key, window_pos):
    global profile_index

    while True:
        with profiles_lock:
            if profile_index >= len(profiles):
                print(f"Thread {thread_id}: No more profiles to process.")
                break
            current_profile_index = profile_index
            profile_index += 1

        profile = profiles[current_profile_index]
        profile_id = profile["id"]
        row_number = profile["row"]

        print(f"Thread {thread_id}: Processing profile {profile_id} (Row {row_number})")

        # Đánh dấu profile đang xử lý
        mark_profile_as_processing(thread_id, profile_id, row_number)

        # 1) Lấy proxy từ API key và cập nhật proxy cho profile
        proxy = None
        while True:
            proxy_result = get_proxy_from_api(api_key)
            
            if isinstance(proxy_result, dict) and "wait" in proxy_result:
                # Cần đợi trước khi lấy proxy mới
                wait_seconds = proxy_result["wait"]
                print(f"Thread {thread_id}: Đợi {wait_seconds} giây trước khi lấy proxy mới.")
                time.sleep(wait_seconds)
                continue
            elif proxy_result:
                # Đã lấy được proxy
                proxy = proxy_result
                break
            else:
                # Lỗi khi lấy proxy, thử lại sau 10 giây
                print(f"Thread {thread_id}: Lỗi khi lấy proxy, thử lại sau 10 giây.")
                time.sleep(10)
                continue
                
        # Cập nhật proxy cho profile
        if not update_proxy(profile_id, proxy):
            print(f"Thread {thread_id}: Skipping profile {profile_id} due to proxy update failure.")
            mark_profile_as_completed(profile_id)
            continue

        # 3) Mở profile qua API
        start_url = f"http://127.0.0.1:19995/api/v3/profiles/start/{profile_id}?addination_args=--lang%3Dvi&win_pos={window_pos}&win_size=1800%2C1080&win_scale=0.35"
        print(f"Thread {thread_id}: Opening profile via URL: {start_url}")
        try:
            start_resp = requests.get(start_url)
            start_resp.raise_for_status()
        except Exception as e:
            print(f"Thread {thread_id}: Error opening profile {profile_id}: {e}")
            mark_profile_as_completed(profile_id)
            continue

        start_data = start_resp.json()
        if not start_data.get("success"):
            print(f"Thread {thread_id}: Failed to open profile {profile_id}: {start_data}")
            mark_profile_as_completed(profile_id)
            continue

        driver_path = start_data.get("data", {}).get("driver_path")
        remote_debugging_address = start_data.get("data", {}).get("remote_debugging_address")
        browser_location = start_data.get("data", {}).get("browser_location")

        if not driver_path or not remote_debugging_address:
            print(f"Thread {thread_id}: Missing driver_path or not remote_debugging_address, skipping profile.")
            mark_profile_as_completed(profile_id)
            continue

        # 4) Khởi tạo Selenium
        options = Options()
        # options.binary_location = browser_location # nếu cần
        options.add_experimental_option("debuggerAddress", remote_debugging_address)
        service = Service(executable_path=driver_path)
        try:
            driver = webdriver.Chrome(service=service, options=options)
            wait = WebDriverWait(driver, 30)
        except Exception as e:
            print(f"Thread {thread_id}: Error initializing webdriver for profile {profile_id}: {e}")
            mark_profile_as_completed(profile_id)
            continue

        try:
            # ---------------------------
            # 5) Truy cập trang đăng ký Outlook
            # ---------------------------
            outlook_signup_url = "https://signup.live.com/signup?mkt=en-US&lic=1"
            print(f"Thread {thread_id}: Navigating to {outlook_signup_url}")
            
            try:
                driver.get(outlook_signup_url)
                wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                print(f"Thread {thread_id}: Outlook signup page loaded.")
                #time.sleep(3)  # Đợi thêm để trang hiển thị hoàn chỉnh
                
                # Tạo username ngẫu nhiên
                first_name, last_name = generate_random_name()
                username_base = (first_name + last_name).lower()
                random_digits = ''.join([str(random.randint(0, 9)) for _ in range(4)])
                username = username_base + random_digits
                email = username + "@outlook.com"  # Thêm @outlook.com trực tiếp vào username
                
                # Nhập username kèm @outlook.com và kiểm tra lỗi
                username_accepted = False
                max_attempts = 5
                attempt = 0
                
                while not username_accepted and attempt < max_attempts:
                    attempt += 1
                    
                    try:
                        username_input = WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.ID, "usernameInput"))
                        )
                        username_input.clear()
                        # Sử dụng simulate_human_typing thay vì send_keys và nhập cả @outlook.com
                        simulate_human_typing(username_input, email)
                        print(f"Thread {thread_id}: Entered email: {email}")
                        
                        # Click nút submit sử dụng simulate_human_click
                        submit_button = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit'][aria-describedby*='usernameCollection']"))
                        )
                        simulate_human_click(driver, submit_button)
                        print(f"Thread {thread_id}: Clicked submit button")
                        
                        # Kiểm tra lỗi trong 5 giây
                        try:
                            error_element = WebDriverWait(driver, 5).until(
                                EC.presence_of_element_located((By.ID, "usernameInputError"))
                            )
                            print(f"Thread {thread_id}: Username error detected. Trying a new username.")
                            
                            # Tạo username mới
                            first_name, last_name = generate_random_name()
                            username_base = (first_name + last_name).lower()
                            random_digits = ''.join([str(random.randint(0, 9)) for _ in range(4)])
                            username = username_base + random_digits
                            email = username + "@outlook.com"  # Cập nhật email với username mới
                        except TimeoutException:
                            # Không tìm thấy lỗi sau 5 giây, có thể username đã được chấp nhận
                            print(f"Thread {thread_id}: Username {email} accepted!")
                            username_accepted = True
                    except Exception as e:
                        print(f"Thread {thread_id}: Error during username input, attempt {attempt}: {e}")
                        time.sleep(2)
                
                if not username_accepted:
                    print(f"Thread {thread_id}: Failed to get a valid username after {max_attempts} attempts")
                    # Đóng profile và tiếp tục với profile tiếp theo
                    close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                    requests.get(close_url)
                    mark_profile_as_completed(profile_id)
                    continue
                
                # Tạo và nhập mật khẩu
                password = generate_secure_password()
                try:
                    password_input = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "Password"))
                    )
                    # Sử dụng simulate_human_typing thay vì send_keys
                    simulate_human_typing(password_input, password)
                    print(f"Thread {thread_id}: Entered password: {password}")
                except TimeoutException:
                    print(f"Thread {thread_id}: Timeout waiting for password input. Could be slow proxy or page error.")
                    # Đánh dấu là lỗi proxy trong Excel
                    with profiles_lock:
                        worksheet.cell(row=row_number, column=10).value = "Lỗi Proxy - Không tìm thấy trường mật khẩu"
                        save_excel_with_retry()
                    
                    # Đóng profile và chuyển đến profile tiếp theo
                    close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                    try:
                        requests.get(close_url)
                        print(f"Thread {thread_id}: Closed profile {profile_id} due to password input timeout")
                    except Exception as e:
                        print(f"Thread {thread_id}: Error closing profile {profile_id}: {e}")
                    
                    mark_profile_as_completed(profile_id)
                    continue
                except Exception as e:
                    print(f"Thread {thread_id}: Error entering password: {e}")
                
                # Click nút next với độ trễ ngẫu nhiên
                try:
                    next_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.ID, "nextButton"))
                    )
                    simulate_human_click(driver, next_button)
                    print(f"Thread {thread_id}: Clicked next button after password")
                except Exception as e:
                    print(f"Thread {thread_id}: Error clicking next button: {e}")
                
                # Nhập họ và tên
                try:
                    last_name_input = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "lastNameInput"))
                    )
                    # Sử dụng simulate_human_typing thay vì send_keys
                    simulate_human_typing(last_name_input, last_name)
                    print(f"Thread {thread_id}: Entered last name: {last_name}")
                    
                    first_name_input = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "firstNameInput"))
                    )
                    # Sử dụng simulate_human_typing thay vì send_keys
                    simulate_human_typing(first_name_input, first_name)
                    print(f"Thread {thread_id}: Entered first name: {first_name}")
                    
                    # Click nút next sử dụng simulate_human_click
                    name_next_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit'][aria-describedby*='signupNameCollection']"))
                    )
                    simulate_human_click(driver, name_next_button)
                    print(f"Thread {thread_id}: Clicked next button after name")
                except Exception as e:
                    print(f"Thread {thread_id}: Error entering name: {e}")
                
                # Nhập ngày tháng năm sinh
                birth_day, birth_month, birth_year = generate_birth_date()

                # Chọn tháng trước để tránh nhầm lẫn
                try:
                    # Click vào dropdown tháng với độ trễ ngẫu nhiên
                    month_select = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.ID, "BirthMonth"))
                    )
                    simulate_human_click(driver, month_select)
                    print(f"Thread {thread_id}: Clicked month select")
                    time.sleep(random.uniform(0.8, 1.5))  # Đợi dropdown hiển thị đầy đủ
                    
                    # Tìm và chọn tháng bằng nhiều cách
                    month_selected = False
                    
                    # Cách 1: Dùng XPath chính xác để định vị option theo vị trí trong dropdown
                    try:
                        month_option = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, f"//select[@id='BirthMonth']/option[@value='{birth_month}']"))
                        )
                        simulate_human_click(driver, month_option)  # Sử dụng click kiểu người dùng
                        month_selected = True
                        print(f"Thread {thread_id}: Selected month: {birth_month} (direct XPath selection)")
                    except Exception as month_error1:
                        print(f"Thread {thread_id}: Could not select month using XPath: {month_error1}")
                    
                    # Cách 2: Sử dụng JavaScript để thiết lập giá trị dropdown
                    if not month_selected:
                        try:
                            driver.execute_script(f"document.getElementById('BirthMonth').value = '{birth_month}';")
                            # Kích hoạt sự kiện change để đảm bảo trình duyệt nhận biết thay đổi
                            driver.execute_script("document.getElementById('BirthMonth').dispatchEvent(new Event('change'));")
                            time.sleep(random.uniform(0.3, 0.7))  # Độ trễ ngẫu nhiên
                            month_selected = True
                            print(f"Thread {thread_id}: Selected month: {birth_month} (using JavaScript)")
                        except Exception as month_error2:
                            print(f"Thread {thread_id}: Could not select month using JavaScript: {month_error2}")
                    
                    # Cách 3: Dùng send_keys với số tháng + độ trễ ngẫu nhiên
                    if not month_selected:
                        try:
                            month_select.clear()  # Xóa giá trị hiện tại nếu có
                            for digit in str(birth_month):
                                month_select.send_keys(digit)
                                time.sleep(random.uniform(0.05, 0.15))  # Độ trễ ngẫu nhiên giữa các ký tự
                            month_select.send_keys(Keys.TAB)  # Ấn Tab để xác nhận lựa chọn
                            time.sleep(random.uniform(0.3, 0.7))
                            month_selected = True
                            print(f"Thread {thread_id}: Selected month: {birth_month} (using send_keys)")
                        except Exception as month_error3:
                            print(f"Thread {thread_id}: Could not select month using send_keys: {month_error3}")
                    
                    if not month_selected:
                        print(f"Thread {thread_id}: Could not select month {birth_month} by any method, using default")
                except Exception as e:
                    print(f"Thread {thread_id}: Error with month dropdown: {e}")

                # Đợi một chút trước khi chọn ngày
                time.sleep(random.uniform(0.8, 1.5))

                # Chọn ngày sau khi đã chọn tháng
                try:
                    # Click vào dropdown ngày với độ trễ ngẫu nhiên
                    day_select = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "BirthDay"))
                    )
                    simulate_human_click(driver, day_select)
                    print(f"Thread {thread_id}: Clicked day select")
                    time.sleep(random.uniform(0.8, 1.5))  # Đợi dropdown hiển thị đầy đủ
                    
                    # Tìm và chọn ngày bằng nhiều cách
                    day_selected = False
                    
                    # Cách 1: Tìm trực tiếp option với value cụ thể
                    try:
                        day_option = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, f"//select[@id='BirthDay']/option[@value='{birth_day}']"))
                        )
                        simulate_human_click(driver, day_option)  # Click kiểu người dùng
                        day_selected = True
                        print(f"Thread {thread_id}: Selected day: {birth_day} (direct selection)")
                    except Exception as day_error1:
                        print(f"Thread {thread_id}: Could not select day directly: {day_error1}")
                    
                    # Cách 2: Nếu không tìm được trực tiếp, thử gửi phím với độ trễ ngẫu nhiên
                    if not day_selected:
                        try:
                            day_select.clear()
                            for digit in str(birth_day):
                                day_select.send_keys(digit)
                                time.sleep(random.uniform(0.05, 0.15))  # Độ trễ ngẫu nhiên giữa các ký tự
                            time.sleep(random.uniform(0.3, 0.7))
                            day_selected = True
                            print(f"Thread {thread_id}: Selected day: {birth_day} (using send_keys)")
                        except Exception as day_error2:
                            print(f"Thread {thread_id}: Could not select day using send_keys: {day_error2}")
                    
                    # Cách 3: Dùng JavaScript để thiết lập giá trị
                    if not day_selected:
                        try:
                            driver.execute_script(f"document.getElementById('BirthDay').value = '{birth_day}';")
                            driver.execute_script("document.getElementById('BirthDay').dispatchEvent(new Event('change'));")
                            time.sleep(random.uniform(0.3, 0.7))
                            day_selected = True
                            print(f"Thread {thread_id}: Selected day: {birth_day} (using JavaScript)")
                        except Exception as day_error3:
                            print(f"Thread {thread_id}: Could not select day using JavaScript: {day_error3}")
                    
                    if not day_selected:
                        print(f"Thread {thread_id}: Could not select day {birth_day} by any method, using default")
                except Exception as e:
                    print(f"Thread {thread_id}: Error with day dropdown: {e}")

                # Nhập năm với độ trễ ngẫu nhiên
                try:
                    year_input = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "BirthYear"))
                    )
                    year_input.clear()
                    # Nhập năm từng ký tự một với độ trễ ngẫu nhiên
                    simulate_human_typing(year_input, str(birth_year))
                    print(f"Thread {thread_id}: Entered birth year: {birth_year}")
                except Exception as e:
                    print(f"Thread {thread_id}: Error entering birth year: {e}")

                # Click nút next với độ trễ ngẫu nhiên
                try:
                    birth_next_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit'][aria-describedby*='countryBirthdateView']"))
                    )
                    simulate_human_click(driver, birth_next_button)
                    print(f"Thread {thread_id}: Clicked next button after birth date")
                except Exception as e:
                    print(f"Thread {thread_id}: Error clicking next button after birth date: {e}")
                
                # Kiểm tra xem tài khoản có bị khóa không sau khi bấm Next
                try:
                    account_locked_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//div[@role='heading' and @aria-level='1' and @id='riskApiBlockedViewTitle']"))
                    )
                    print(f"Thread {thread_id}: [CẢNH BÁO] Phát hiện thông báo tài khoản bị khóa")
                    
                    
                    # Đánh dấu là tài khoản bị khóa trong Excel
                    with profiles_lock:
                        worksheet.cell(row=row_number, column=10).value = "Tài khoản bị khóa"
                        # Tô đỏ dòng đó
                        for col in range(1, 12):  # Bao gồm cả cột K (cột thứ 11)
                            worksheet.cell(row=row_number, column=col).fill = red_fill
                        # Lưu file Excel
                        save_excel_with_retry()
                        print(f"Thread {thread_id}: Đã đánh dấu 'Tài khoản bị khóa' cho profile {profile_id}")
                    
                    # Đóng profile
                    close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                    try:
                        requests.get(close_url)
                        print(f"Thread {thread_id}: Đã đóng profile {profile_id} do tài khoản bị khóa")
                    except Exception as e:
                        print(f"Thread {thread_id}: Lỗi khi đóng profile {profile_id}: {e}")
                    
                    return "ACCOUNT_LOCKED"
                    
                except TimeoutException:
                    print(f"Thread {thread_id}: Không phát hiện thông báo tài khoản bị khóa sau khi bấm Next")
                except Exception as e:
                    print(f"Thread {thread_id}: Lỗi khi kiểm tra tài khoản bị khóa trước bước newSessionLink: {e}")

                # Kiểm tra và đợi animation loading
                try:
                    # Tìm phần tử animation loading
                    loading_element = WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div[data-testid='routeAnimation']"))
                    )
                    print(f"Thread {thread_id}: Found loading animation, waiting for it to disappear (max 60s)")
                    
                    # Đợi cho animation biến mất
                    animation_disappeared = False
                    start_time = time.time()
                    while time.time() - start_time < 120:
                        try:
                            # Kiểm tra xem element còn hiển thị không
                            if not loading_element.is_displayed():
                                animation_disappeared = True
                                print(f"Thread {thread_id}: Loading animation disappeared after {int(time.time() - start_time)} seconds")
                                break
                            
                            # Thử kiểm tra bằng JavaScript
                            is_visible = driver.execute_script(
                                "return arguments[0].offsetParent !== null && "
                                "arguments[0].offsetWidth > 0 && "
                                "arguments[0].offsetHeight > 0 && "
                                "window.getComputedStyle(arguments[0]).visibility !== 'hidden'", 
                                loading_element
                            )
                            
                            if not is_visible:
                                animation_disappeared = True
                                print(f"Thread {thread_id}: Loading animation not visible after {int(time.time() - start_time)} seconds (JS check)")
                                break
                                
                        except Exception:
                            # Nếu element không còn trong DOM, coi như đã biến mất
                            animation_disappeared = True
                            print(f"Thread {thread_id}: Loading animation element no longer in DOM after {int(time.time() - start_time)} seconds")
                            break
                            
                        # Đợi 1 giây trước khi kiểm tra lại
                        time.sleep(1)
                    
                    # Nếu sau 60 giây animation vẫn còn, đóng profile và chuyển đến profile tiếp theo
                    if not animation_disappeared:
                        print(f"Thread {thread_id}: Loading animation still visible after 60 seconds, marking as error and closing profile")
                        
                        # Đánh dấu là lỗi giải captcha trong Excel
                        with profiles_lock:
                            worksheet.cell(row=row_number, column=10).value = "Lỗi Giải Captcha"
                            # Tô đỏ dòng đó
                            for col in range(1, 11):
                                worksheet.cell(row=row_number, column=col).fill = red_fill
                            # Lưu file Excel
                            save_excel_with_retry()
                            print(f"Thread {thread_id}: Đã đánh dấu Lỗi Giải Captcha cho profile {profile_id}")
                        
                        # Đóng profile
                        close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
                        try:
                            requests.get(close_url)
                            print(f"Thread {thread_id}: Closed profile {profile_id} due to captcha timeout")
                        except Exception as e:
                            print(f"Thread {thread_id}: Error closing profile {profile_id}: {e}")
                            
                        # Chuyển đến profile tiếp theo
                        mark_profile_as_completed(profile_id)
                        continue
                    
                except TimeoutException:
                    print(f"Thread {thread_id}: No loading animation detected")
                except NoSuchElementException:
                    print(f"Thread {thread_id}: No loading animation found")
                except Exception as e:
                    print(f"Thread {thread_id}: Error handling loading animation: {e}")
                
                # Click nút OK (nếu xuất hiện)
                try:
                    ok_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "span.ms-Button-label[id='id__0']"))
                    )
                    ok_button.click()
                    print(f"Thread {thread_id}: Clicked OK button")
                except TimeoutException:
                    print(f"Thread {thread_id}: OK button not found or not clickable")
                except Exception as e:
                    print(f"Thread {thread_id}: Error clicking OK button: {e}")
                
                # Thêm các bước mở rộng
                # -------------------------------------------------
                # 1. Click vào nút Yes nếu xuất hiện và đợi trang tải xong
                try:
                    yes_button = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[text() = 'Yes']"))
                    )
                    simulate_human_click(driver, yes_button)
                    print(f"Thread {thread_id}: Clicked Yes button")
                    # Đợi trang load xong
                    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                    time.sleep(3)
                except TimeoutException:
                    print(f"Thread {thread_id}: Yes button not found or not clickable (might not appear)")
                except Exception as e:
                    print(f"Thread {thread_id}: Error clicking Yes button: {e}")
                
                # 2. Truy cập trang bảo mật
                try:
                    security_url = "https://account.microsoft.com/security?lang=en-US#main-content-landing-react"
                    print(f"Thread {thread_id}: Navigating to security page: {security_url}")
                    driver.get(security_url)
                    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                    print(f"Thread {thread_id}: Security page loaded")
                    #time.sleep(3)  # Đợi thêm để đảm bảo trang đã tải hoàn tất
                except Exception as e:
                    print(f"Thread {thread_id}: Error accessing security page: {e}")
                
                # 3. Click vào 'View my sign-in activity'
                try:
                    view_signin_link = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='View my sign-in activity']"))
                    )
                    simulate_human_click(driver, view_signin_link)
                    print(f"Thread {thread_id}: Clicked 'View my sign-in activity'")
                    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                    #time.sleep(5)  # Đợi thêm để trang tải hoàn tất
                except TimeoutException:
                    print(f"Thread {thread_id}: 'View my sign-in activity' link not found")
                except Exception as e:
                    print(f"Thread {thread_id}: Error clicking 'View my sign-in activity': {e}")
                
                # 4. Chọn ngẫu nhiên một Gmail từ file recovery_mail.txt
                recovery_email_data = None
                max_retry_email = 10
                retry_count = 0
                
                while not recovery_email_data and retry_count < max_retry_email:
                    retry_count += 1
                    recovery_email_data = get_recovery_email(thread_id)
                    if not recovery_email_data:
                        print(f"Thread {thread_id}: Không thể lấy email khôi phục, đợi 5 giây và thử lại ({retry_count}/{max_retry_email})")
                        time.sleep(5)
                
                if not recovery_email_data:
                    print(f"Thread {thread_id}: Không thể lấy email khôi phục sau {max_retry_email} lần thử, tiếp tục quy trình")
                    # Đóng profile và chuyển đến profile tiếp theo
                    release_recovery_email(thread_id)  # Đảm bảo giải phóng email nếu đã được giữ
                    mark_profile_as_completed(profile_id)
                    continue
                
                gmail, client_id, client_secret, gmail_refresh_token = recovery_email_data
                print(f"Thread {thread_id}: Đã chọn Gmail {gmail} để sử dụng làm email khôi phục")
                
                # 5. Nhập Gmail vào trường email khôi phục
                try:
                    email_input = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@id='EmailAddress']"))
                    )
                    simulate_human_typing(email_input, gmail)
                    print(f"Thread {thread_id}: Đã nhập email khôi phục: {gmail}")
                    
                    # Click nút tiếp theo
                    next_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='iNext']"))
                    )
                    simulate_human_click(driver, next_button)
                    print(f"Thread {thread_id}: Clicked Next button after entering recovery email")
                    
                    # Đợi trang tải xong
                    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                    time.sleep(5)  # Đợi thêm để trang hoàn tất và email xác minh được gửi
                except TimeoutException:
                    print(f"Thread {thread_id}: Email input or Next button not found")
                    release_recovery_email(thread_id)
                except Exception as e:
                    print(f"Thread {thread_id}: Error entering recovery email: {e}")
                    release_recovery_email(thread_id)
                
                # 6. Đọc email để lấy mã xác minh
                verification_code = None
                max_retries_code = 5
                
                for retry_code in range(max_retries_code):
                    print(f"Thread {thread_id}: Đang thử lấy mã xác minh, lần thử {retry_code + 1}/{max_retries_code}")
                    verification_code = get_verification_code_from_gmail(
                        thread_id, 
                        email,  # Email Outlook vừa đăng ký
                        gmail,  # Email Gmail đang sử dụng để xác minh
                        client_id, 
                        client_secret, 
                        gmail_refresh_token
                    )
                    
                    if verification_code:
                        print(f"Thread {thread_id}: Đã lấy được mã xác minh: {verification_code}")
                        break
                    
                    print(f"Thread {thread_id}: Chưa tìm thấy mã xác minh, đợi 5 giây và thử lại")
                    time.sleep(5)
                
                if not verification_code:
                    print(f"Thread {thread_id}: Không thể lấy được mã xác minh sau {max_retries_code} lần thử")
                    # Đánh dấu là thất bại trong việc xác minh email
                    with profiles_lock:
                        worksheet.cell(row=row_number, column=10).value = "Lỗi Xác Minh Email"
                        save_excel_with_retry()
                    
                    release_recovery_email(thread_id)
                    continue
                
                # 7. Nhập mã xác minh vào trường iOttText
                try:
                    code_input = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@id='iOttText']"))
                    )
                    simulate_human_typing(code_input, verification_code)
                    print(f"Thread {thread_id}: Đã nhập mã xác minh: {verification_code}")
                    
                    # Click nút xác nhận
                    verify_button = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@id='iNext']"))
                    )
                    simulate_human_click(driver, verify_button)
                    print(f"Thread {thread_id}: Clicked Next button after entering verification code")
                    
                    # Đợi trang tải xong
                    time.sleep(10)
                    
                    # Lưu email khôi phục vào Cột K (cột thứ 11)
                    with profiles_lock:
                        worksheet.cell(row=row_number, column=11).value = gmail
                        save_excel_with_retry()
                        print(f"Thread {thread_id}: Đã lưu email khôi phục {gmail} vào cột K")
                    
                except TimeoutException:
                    print(f"Thread {thread_id}: Verification code input or button not found")
                except Exception as e:
                    print(f"Thread {thread_id}: Error entering verification code: {e}")
                
                # Giải phóng email khôi phục để các thread khác có thể sử dụng
                release_recovery_email(thread_id)
                
                # Đợi một chút cho tài khoản được thiết lập hoàn toàn
                #time.sleep(5)
                
                # Truy cập https://tolive.site để lấy refresh token
                print(f"Thread {thread_id}: Accessing tolive.site to get refresh token")
                refresh_token = get_refresh_token(driver, wait, username, password, thread_id, profile_id, row_number)

                if refresh_token == "FAILED" or refresh_token == "FAILED_ANIMATION":
                    print(f"Thread {thread_id}: Đánh dấu tài khoản này là Thất Bại")
                    
                    # Lưu kết quả vào Excel
                    with profiles_lock:
                        # Đánh dấu thất bại vào cột J (cột thứ 10)
                        worksheet.cell(row=row_number, column=10).value = "Thất Bại"
                        
                        # Lưu file Excel
                        save_excel_with_retry()
                        print(f"Thread {thread_id}: Đã đánh dấu Thất Bại cho profile {profile_id}")
                elif refresh_token:
                    print(f"Thread {thread_id}: Successfully got refresh token")
                    
                    # Lưu thông tin vào file Excel
                    with profiles_lock:
                        # Lưu email vào cột G (cột thứ 7)
                        worksheet.cell(row=row_number, column=7).value = f"{username}@outlook.com"
                        
                        # Lưu mật khẩu vào cột H (cột thứ 8)
                        worksheet.cell(row=row_number, column=8).value = password
                        
                        # Lưu refresh token vào cột I (cột thứ 9)
                        worksheet.cell(row=row_number, column=9).value = refresh_token
                        
                        # Đánh dấu thành công vào cột J (cột thứ 10)
                        worksheet.cell(row=row_number, column=10).value = "Thành Công"
                        
                        # Đặt màu dòng thành mặc định (loại bỏ màu đỏ nếu có)
                        for col in range(1, 11):
                            worksheet.cell(row=row_number, column=col).fill = default_fill
                        
                        # Lưu file Excel
                        save_excel_with_retry()
                        print(f"Thread {thread_id}: Saved account info to Excel for profile {profile_id} and removed red highlight")
                else:
                    print(f"Thread {thread_id}: Failed to get refresh token")
                    
                    # Lưu kết quả vào Excel - đánh dấu là Lỗi
                    with profiles_lock:
                        worksheet.cell(row=row_number, column=10).value = "Lỗi Token"
                        save_excel_with_retry()
                        print(f"Thread {thread_id}: Đã đánh dấu Lỗi Token cho profile {profile_id}")
            except Exception as e:
                print(f"Thread {thread_id}: Error during Outlook registration process: {e}")
                # Đảm bảo đánh dấu lỗi vào cột J
                with profiles_lock:
                    worksheet.cell(row=row_number, column=10).value = f"Lỗi: {str(e)[:50]}"
                    save_excel_with_retry()
        finally:
            # Đóng profile qua API
            close_url = f"http://127.0.0.1:19995/api/v3/profiles/close/{profile_id}"
            print(f"Thread {thread_id}: Closing profile with URL: {close_url}")
            try:
                requests.get(close_url)
                print(f"Thread {thread_id}: Profile {profile_id} closed successfully.")
            except Exception as e:
                print(f"Thread {thread_id}: Error closing profile {profile_id}: {e}")
            
            mark_profile_as_completed(profile_id)
            time.sleep(1)

# ---------------------------------------
# Main
# ---------------------------------------
if __name__ == "__main__":
    # Khôi phục trạng thái từ lần chạy trước
    recover_from_previous_run()

    # Xác định số lượng luồng dựa trên số API key có sẵn
    num_threads = min(7, len(api_keys))
    
    if num_threads == 0:
        print("Không có API key nào trong file proxy.txt. Không thể chạy chương trình.")
        exit(1)
    
    # Tạo và chạy các luồng dựa trên số API key có sẵn
    threads = []
    window_positions = ["0,0", "1800,0", "3600,0", "0,1080", "1800,1080", "3600,1080", "0,2160"]
    
    for i in range(num_threads):
        thread = threading.Thread(
            target=process_profile, 
            args=(i+1, api_keys[i], window_positions[i])
        )
        threads.append(thread)
        thread.start()
        
    # Chờ tất cả các luồng hoàn thành
    for thread in threads:
        thread.join()

    print("\nCompleted processing all profiles.")
